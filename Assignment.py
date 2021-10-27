import pandas as pd
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
try:
    os.remove("EP1.xlsx")
except Exception as e:
    print("OK")


headers       = ['ID','NAME','DOJ']
workbook_name = 'EP1.xlsx'
wb = Workbook()
page = wb.active
page.title = 'emp'
page.append(headers) # write the headers to the first line
companies = [['name1','address1','tel1'], ['name2','address2','tel2']]
for info in companies:
    page.append(info)
wb.save(filename = workbook_name)

while 1==1:
    print("=========Menu==========\n\n1.View Records\n2.Add New Records\n3.Update Records\n4.Delete Record\n5.Quit")
    value_selected=input("Please select option from above Menu:")
    if int(value_selected) == 5:
        print("Stopping Programme")
        break
    elif int(value_selected) == 1:
        df = pd.read_excel('EP1.xlsx', engine='openpyxl')
        print(df.to_string(index=False))
        continue
    elif int(value_selected) == 2:
        data = []
        for val in ("ID","NAME","DOJ"):
            value=input("Please Provide %s:"%val)
            data.append(value)
        workbook_name = 'EP1.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
        page.append(data)
        wb.save(filename=workbook_name)
        continue
    elif int(value_selected) == 3:
        df = pd.read_excel('EP1.xlsx', engine='openpyxl')
        print(df)
        row=input("Please provide row number to update:")
        for col in ("ID","NAME","DOJ"):
            val=input("Please provide New %s Value(Press Enter to Skip update of %s Column):"%(col,col))
            if val == "":
                continue
            else:
                df['%s'%str(col)][int(row)] = val
        print(df.to_string(index=False))
        writer = pd.ExcelWriter('EP1.xlsx', engine='openpyxl')
        df.to_excel(writer, sheet_name = 'emp',index= False)
        writer.save()
        writer.close()
        continue
    elif int(value_selected) == 4:
        from openpyxl import load_workbook
        workbook_name = 'EP1.xlsx'
        wb = load_workbook(workbook_name)
        sheet = wb.active
        sheet.delete_rows(2,1)
        wb.save(filename=workbook_name)
        df = pd.read_excel('EP1.xlsx', engine='openpyxl')
        print(df)
        continue
    else:
        print("Please select valid option to continue")
        continue
